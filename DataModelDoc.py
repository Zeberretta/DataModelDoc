######  IMPORTS
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from datetime import datetime as dt
import pytz

import warnings
warnings.filterwarnings("ignore")

######  FUNCTIONS


# Function to adjust width of all columns in the workbook and add the homepage button for table sheets
def adjust_col_width(sheet):
    # Add Homepage Button
    if sheet.title not in ['InfoPage', 'Navigation', 'Subject Areas']:
        sheet.cell(row=1, column=1).value = '⌂'
        sheet.cell(row=1, column=1).hyperlink = '#\'Navigation\'!A1'
    # Adjust each column width
    for col in sheet.columns:
         max_length = 0
         column = col[0].column_letter # Get the column name
         for cell in col:
             try: # Necessary to avoid error on empty cells
                 if len(str(cell.value)) > max_length:
                     max_length = len(str(cell.value))
             except:
                 pass
         adjusted_width = (max_length + 2) * 1.2
         sheet.column_dimensions[column].width = adjusted_width
    sheet.column_dimensions[[col[0].column_letter for col in sheet.columns][0]].width = 3 # Column 'A' has a fixed width

# Function do set border around a range - it must have at least 2 col and 2 rows
def set_border(ws, cell_range):
    Border(left=Side(border_style='thin', color='000000'))
    rows = ws[cell_range]
    # First row
    for cell in rows[0]:
        cell.border = Border(top=Side(border_style='thin', color='000000'))
    # Last row
    for cell in rows[-1]:
        cell.border = Border(bottom=Side(border_style='thin', color='000000'))
    # First col
    for cell in rows:
        cell[0].border = Border(left=Side(border_style='thin', color='000000'))
    # Last col
    for cell in rows:
        cell[-1].border = Border(right=Side(border_style='thin', color='000000'))
    # Corners
    rows[0][0].border = Border(top=Side(border_style='thin', color='000000'),
                                    left=Side(border_style='thin', color='000000'))
    rows[0][-1].border = Border(top=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'))
    rows[-1][0].border = Border(bottom=Side(border_style='thin', color='000000'),
                                    left=Side(border_style='thin', color='000000'))
    rows[-1][-1].border = Border(bottom=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'))

# Funtction to format each sheet with the same standard
def format_sheet(ws):

    # Setting sheet title and its formatting
    ws['B7'].alignment = Alignment(horizontal='center')
    ws['B7'].value = ws.title
    ws['B7'].font = Font(name='Trebuchet MS', bold=True, size=14)
    ws['B8'].alignment = Alignment(horizontal='center')
    ws['B8'].value = 'Do not change Tab name, modify column order, or sort!'
    #Upper color bar
    mxDim_col = ws.max_column
    mxDim_row = ws.max_row
    for i in range(2, 10):
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=mxDim_col+1)
    ws.row_dimensions[2].height = 5
    ws.row_dimensions[3].height = 7
    ws.row_dimensions[4].height = 9
    ws['B2'].fill = PatternFill(start_color="001F497D", fill_type="solid")
    ws['B3'].fill = PatternFill(start_color="00538DD5", fill_type="solid")
    ws['B4'].fill = PatternFill(start_color="00C5D9F1", fill_type="solid")

    #Setting every cell to have white background
    for cell_line in ws['A1:S100']:
        for cell in cell_line:
            cell.fill = PatternFill(start_color="00FFFFFF", fill_type="solid")


    # Table title color
    for cell in ws['C11:{}11'.format(chr(mxDim_col+64))][0]:
        # print(cell)
        cell.fill = PatternFill(start_color="001F497D", end_color="001F497D", fill_type="solid")

    # Set border around the table and the title
    set_border(ws, 'B2:{}{}'.format(chr(mxDim_col+64+1), mxDim_row+1))
    set_border(ws, 'B5:{}{}'.format(chr(mxDim_col+64+1), mxDim_row+1))

    # Set border around each cell of the table
    rows = ws['C11:{}{}'.format(chr(mxDim_col+64), mxDim_row)]
    for row in rows:
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'),
                                 left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'))

    # For the Navigation sheet, centralize the 'X's cells
    if 'Navigation' in ws.title:
        rows = ws['G12:{}{}'.format(chr(mxDim_col + 64), mxDim_row)]
        for row in rows:
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

# Function to add hyperlinks on cells to make easier to navigate
def set_hyperlinks(ws):
    for i in range(2, ws.max_row - 9):
        name = ws.cell(row=i + 10, column=1 + 2).value
        # print(ws.title)
        try:
            if len(name) > 31:  name = name.replace(' ', '').replace('-', ' -')
        except:
            print(name)
        link = '#\'' + name + '\'!A1'
        # print(link)
        ws.cell(row=i + 10, column=1 + 2).hyperlink = link


# Mais function to extract data from CSV/XML and turn it in the formatted workbook
def Data_Process(path_source, path_base, types_path, name):
    ######  INITIALIZATIONS
    # Date
    CurrentDate = dt.now(pytz.timezone('US/Eastern')).strftime("%m/%d/%Y %I:%M %p")
    #Get Types Dict
    df = pd.read_xml(types_path, xpath="//COLUMN")
    df = df.drop_duplicates(subset=['NAME'])
    for i_type in df.index:
        if pd.notnull(df['LENGTH'][i_type]):
            #print(i_type)
            df['TYPE'][i_type] = df['TYPE'][i_type]+'(' + str(int(df['LENGTH'][i_type])) + ')'
    typeDict = pd.Series(df['TYPE'].values, index=df['NAME']).to_dict()

    #Set some dict
    PurpDict = {'Dim': 'Dimension', 'Fact': 'Fact', 'Hier': 'Hierarchy'} #Porpouses dict
    TabsDict = {'Dim': '#B8CCE4', 'Fact': '#FCD5B4', 'Hier': '#CCC0DA'
                , 'InfoPage': '#1F497D', 'Navigation': '#92D050', 'Subject Areas': '#1F497D', 'Analytics':'#00B0F0'} #Color dict

    #Read source info
    file = pd.read_csv(path_source)

    #List all Subject Areas
    SubAreas = file[file.columns[0]].unique()
    SubAreas = SubAreas[~pd.isnull(SubAreas)] #Remove NaN value

    #List all Tables
    TblNames = file[file.columns[1]].unique()
    TblNames = TblNames[~pd.isnull(TblNames)] #Remove NaN value
    TblNames = sorted(TblNames, reverse=True) #Sort names

    #Open new excel
    writer = pd.ExcelWriter(path_base, engine= 'xlsxwriter')

    ### WRITING THE CONTENTS

    #InfoPage Sheet
    header_Info = ['Name', 'Info']
    newDF = [pd.DataFrame(columns=header_Info)]
    newDF[-1]['Name'] = ['Company', 'Client', 'Project', 'Created by', 'Date', 'Compartment', 'Database', 'OCID']
    newDF[-1]['Info'] = ['Peloton Group, LLC', 'Liberty Oil Field Services', 'Liberty Oil Field Services ADW'] + [name, CurrentDate] + ['None', 'None', '']
    newDF[-1].to_excel(writer,
                 sheet_name='InfoPage', index=None, startrow=11, startcol=2, header=None)
    writer.sheets['InfoPage'].set_tab_color(TabsDict['InfoPage'])

    #Navigation Sheet
    header_N = ['Object Name', 'Type', 'Purpose', 'Schema Name'] + list(SubAreas) + ['Object Comment']
    newDF.append(pd.DataFrame(columns=header_N))
    newDF[-1]['Object Name'] = TblNames
    newDF[-1]['Type'] = 'Table'  # CHECK THIS INFO
    newDF[-1]['Purpose'] = [PurpDict[tbl.split(' - ')[0]] for tbl in TblNames]

    for tbl in TblNames:
        lista = file[file[file.columns[1]] == tbl]['Physical Schema'].unique()
        lista = lista[~pd.isnull(lista)]
        newDF[-1]['Schema Name'] = lista[0]
    for sub in SubAreas:
        for i_tbl in range(len(TblNames)):
            if TblNames[i_tbl] in file[file[file.columns[0]] == sub]['Presentation Table'].unique():
                newDF[-1][sub][i_tbl] = 'x'
            else:
                newDF[-1][sub][i_tbl] = ''
    newDF[-1].to_excel(writer,
                 sheet_name='Navigation', index=None, startrow=10, startcol=2)
    writer.sheets['Navigation'].add_table(10, 2, len(TblNames)+10, len(SubAreas)+6,  {'columns': [{'header': column} for column in header_N]})
    writer.sheets['Navigation'].set_tab_color(TabsDict['Navigation'])

    #Subject Area Sheet
    header_SA = ['Object Name', 'Type', 'Purpose', 'Object Comment']
    newDF.append(pd.DataFrame(columns=header_SA))
    newDF[-1]['Object Name'] = SubAreas
    newDF[-1]['Type'] = 'Subject Area'
    newDF[-1].to_excel(writer,
                 sheet_name='Subject Areas', index=None, startrow=10, startcol=2)
    writer.sheets['Subject Areas'].set_tab_color(TabsDict['Subject Areas'])
    writer.sheets['Subject Areas'].add_table(10, 2, len(SubAreas)+10, 5,  {'columns': [{'header': column} for column in header_SA]})

    #Each Subject Area Sheet
    header_1 = ['Object Name', 'Type', 'Purpose', 'Schema Name', 'Object Comment']
    for sub in SubAreas:
        newDF.append(pd.DataFrame(columns=header_1))
        filt_file = file[file[file.columns[0]] == sub]
        PresTbl = filt_file[filt_file.columns[1]].unique()
        newDF[-1]['Object Name'] = PresTbl
        newDF[-1]['Type'] = 'Table'  # CHECK THIS INFO
        newDF[-1]['Purpose'] = [PurpDict[tbl.split(' - ')[0]] for tbl in PresTbl]
        filt_file = filt_file[~pd.isnull(filt_file['Physical Schema'])] # Remove rows with null Schemas
        newDF[-1]['Schema Name'] = [list(filt_file[filt_file[filt_file.columns[1]] == tbl]['Physical Schema'].unique())[0] for
                                    tbl in PresTbl]
        newDF[-1].to_excel(writer, sheet_name=sub, index=None, startrow=10, startcol=2)
        writer.sheets[sub].set_tab_color(TabsDict['Analytics'])
        writer.sheets[sub].add_table(10, 2, len(PresTbl)+10, 6,  {'columns': [{'header': column} for column in header_1]})


    #Table Sheets
    header_2 = ['OAC Display Name', 'Database Field Name', 'Field Data Type', 'Column Availability OAC',
                'Database Field Description', 'Database Table Name', 'Transformation Rule', 'Expression']
    for sub in TblNames:
        #print(sub)
        newDF.append(pd.DataFrame(columns=header_2))
        filt_file = file[file[file.columns[1]] == sub]
        #PresTbl = filt_file[filt_file.columns[2]].unique()
        newDF[-1][header_2[0]] = filt_file['Presentation Column']
        newDF[-1][header_2[1]] = filt_file['Physical Column']
        newDF[-1][header_2[-1]] = filt_file['Expression']
        newDF[-1][header_2[2]] = [ typeDict[x] for x in list(filt_file['Logical Column'])]
        if len(sub) > 31:
            sub = sub.replace(' ','').replace('-',' -')
        newDF[-1].to_excel(writer, sheet_name=sub, index=None, startrow=1, startcol=1)
        writer.sheets[sub].set_tab_color(TabsDict[sub.split(' -')[0]])
        writer.sheets[sub].add_table(1, 1, len(filt_file['Presentation Column']) + 1, 8, {'columns': [{'header': column} for column in header_2]})
        #writer.sheets[sub].ce


    writer.save()
    writer.close()

    ### FORMATTING

    wb = load_workbook(path_base) #open it up again to format
    # Default formatting for the first sheets
    for ws in wb._sheets:
        adjust_col_width(ws)
        if 'InfoPage' not in ws.title and ' -' not in ws.title:
            # print(ws.title)
            set_hyperlinks(ws)  # HYPERLINKS
        if ' -' not in ws.title:
            format_sheet(ws)


    wb.save(path_base)
    wb.close()

if __name__ == '__main__':
    # Set file paths
    path_source = r'D:\Users\Jose\Documents\TRI\Liberty Oilfields\Repo_Doc.csv'
    path_base = r'D:\Users\Jose\Documents\TRI\Liberty Oilfields\Data_Model_Documenation_LOS_new.xlsx'
    types_path = r'D:\Users\Jose\Documents\TRI\Liberty Oilfields\logicalType.xml'
    #Call main function
    Data_Process(path_source, path_base, types_path, 'José Berretta')